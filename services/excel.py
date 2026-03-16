import io
import requests
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# 圖片尺寸（像素）
IMG_PX = 100
# 對應 Excel 列高（點）與欄寬（字元）
IMG_ROW_HEIGHT = 78
IMG_COL_WIDTH  = 14

HEADERS    = ["排名", "圖片", "商品名稱", "價格(NT$)", "銷量", "商品連結"]
COL_WIDTHS = [6, IMG_COL_WIDTH, 42, 12, 10, 50]

HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

REFERER_MAP = {
    "momoshop.com.tw":     "https://www.momoshop.com.tw/",
    "ikea.com.tw":         "https://www.ikea.com.tw/",
    "trplus.com.tw":       "https://www.trplus.com.tw/",
    "pchome.com.tw":       "https://24h.pchome.com.tw/",
    "img.pchome.com.tw":   "https://24h.pchome.com.tw/",
}

def _get_referer(url):
    for domain, referer in REFERER_MAP.items():
        if domain in url:
            return referer
    return "https://www.google.com/"

def _download_image(url):
    """下載並縮圖，回傳 PNG BytesIO；失敗回傳 None"""
    if not url:
        return None
    try:
        r = requests.get(url, timeout=8, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": _get_referer(url),
        })
        r.raise_for_status()
        pil_img = PILImage.open(io.BytesIO(r.content)).convert("RGBA")
        pil_img.thumbnail((IMG_PX, IMG_PX), PILImage.LANCZOS)
        out = io.BytesIO()
        pil_img.save(out, format="PNG")
        out.seek(0)
        return out
    except Exception:
        return None

def _style_header(ws):
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def export(keyword, results_by_platform):
    # 並行下載所有商品圖片
    all_urls = list({
        p.get("image_url", "")
        for products in results_by_platform.values()
        for p in products
        if p.get("image_url")
    })
    with ThreadPoolExecutor(max_workers=10) as executor:
        img_cache = dict(zip(all_urls, executor.map(_download_image, all_urls)))

    wb = Workbook()
    wb.remove(wb.active)  # 移除預設空白工作表

    for platform, products in results_by_platform.items():
        ws = wb.create_sheet(title=platform)
        _style_header(ws)

        for row_idx, p in enumerate(products, start=2):
            values = [
                p.get("rank", ""),
                "",                     # 圖片欄留空，由 add_image 填入
                p.get("name", ""),
                p.get("price", ""),
                p.get("sales", "N/A"),
                p.get("product_url", ""),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 6)))

            # 商品連結設為超連結
            url = p.get("product_url", "")
            if url:
                link_cell = ws.cell(row=row_idx, column=6)
                link_cell.hyperlink = url
                link_cell.font = Font(color="0563C1", underline="single")

            # 嵌入商品圖片
            img_data = img_cache.get(p.get("image_url", ""))
            if img_data:
                xl_img = XLImage(img_data)
                xl_img.width  = IMG_PX
                xl_img.height = IMG_PX
                ws.add_image(xl_img, f"B{row_idx}")

            ws.row_dimensions[row_idx].height = IMG_ROW_HEIGHT

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
